"""Output Format Handlers"""
import json
import csv
import logging
from pathlib import Path
from collections import Counter
from jinja2 import Environment, FileSystemLoader, select_autoescape
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# Canonical column definitions: (Internal Key, Display Name)
REPORT_COLUMNS = [
    ("component_name", "Component Name"),
    ("id", "Vuln ID"),
    ("severity", "Severity"),
    ("cvssScore", "Score"),
    ("component_version", "Library Version"),
    ("upgrade_to", "Upgrade To"),
    ("delta_status", "Delta Status"),
    ("path", "Path"),
    ("published", "Published"),
    ("description", "Description"),
    ("feed_file", "Feed")
]

def get_report_columns(has_baseline=False):
    """Retrieve active report columns, omitting Delta Status if baseline is not active."""
    if has_baseline:
        return REPORT_COLUMNS
    return [col for col in REPORT_COLUMNS if col[0] != "delta_status"]

def is_ui_finding(row):
    """Determine if a vulnerability finding belongs to a UI/frontend library."""
    feed = str(row.get("feed_file") or "").upper()
    tool = str(row.get("tool") or "").upper()
    purl = str(row.get("purl") or "").upper()
    path = str(row.get("path") or "").upper()
    comp_name = str(row.get("component_name") or "").upper()
    
    # Check if npm, retire.js, or javascript-related
    if "RETIRE" in tool or "NPM" in tool:
        return True
    if "PKG:NPM" in purl or "OSV - NPM" in feed or "NPM AUDIT" in tool:
        return True
    if "WEB-UI" in path or "TENANT-ADMIN-WEB-UI" in path or "SPLASHAI-WEB-UI" in path:
        return True
    if "NODE_MODULES" in path or "/JS/" in path or path.endswith(".JS"):
        return True
    if "JQUERY" in comp_name or "DOMPURIFY" in comp_name or "BOOTSTRAP" in comp_name:
        return True
    return False

def categorize_findings(processed_data):
    """Categorize findings into mutually exclusive groups based on tool/format.
    Only categories with > 0 items are included in the returned dict to avoid empty tabs/sheets.
    """
    all_categories = {
        "Jars / War / Zip": [],
        "Npm audit": [],
        "RetireJS": [],
        "Pip-audit": [],
        "OS & System Packages": []
    }
    
    for row in processed_data:
        pkg_type = str(row.get("pkg_type") or "").lower()
        tool = str(row.get("tool") or "").lower()
        path = str(row.get("path") or "").lower()
        feed = str(row.get("feed_file") or "").lower()
        purl = str(row.get("purl") or "").lower()
        comp_name = str(row.get("component_name") or "").lower()
        
        # 1. RetireJS
        if "retire" in tool or "retire" in feed or "retire" in comp_name:
            all_categories["RetireJS"].append(row)
        # 2. Npm audit
        elif (
            pkg_type == "npm_package" or
            "npm audit" in tool or "npm-audit" in tool or "npm" in tool or "npm" in feed or 
            "pkg:npm" in purl or "node_modules" in path or "package.json" in path
        ):
            all_categories["Npm audit"].append(row)
        # 3. Pip-audit
        elif (
            pkg_type == "python_package" or
            "pip-audit" in tool or "pip_audit" in tool or "pip" in tool or "python" in tool or "pypi" in feed or
            "pkg:pypi" in purl or "site-packages" in path or "requirements.txt" in path or "pyproject.toml" in path
        ):
            all_categories["Pip-audit"].append(row)
        # 4. Jars / War / Zip
        elif (
            pkg_type == "java_package" or
            "java" in purl or "maven" in purl or 
            "pkg:maven" in purl or "pkg:java" in purl or
            any(ext in path for ext in (".jar", ".war", ".zip", ".ear")) or
            "boot-inf" in path or "meta-inf" in path
        ):
            all_categories["Jars / War / Zip"].append(row)
        # 5. OS & System Packages (Fallback)
        else:
            all_categories["OS & System Packages"].append(row)
            
    # Filter out empty categories to ensure sheets/tabs are only created if they contain findings
    return {k: v for k, v in all_categories.items() if len(v) > 0}

def calculate_cvss_score(vector_str):
    """Calculates CVSS v3.x or CVSS v4.0 scores from their vector strings in pure Python."""
    if isinstance(vector_str, (int, float)):
        return f"{float(vector_str):.1f}"
    if not isinstance(vector_str, str):
        return "N/A"
    vector_str = vector_str.strip()
    if not vector_str.startswith("CVSS:"):
        try:
            val = float(vector_str)
            return f"{val:.1f}"
        except ValueError:
            return vector_str
            
    # Parse vector into a dict
    parts = vector_str.split('/')
    metrics = {}
    for p in parts:
        if ':' in p:
            k, v = p.split(':', 1)
            metrics[k] = v
            
    version = "3.1"
    if parts[0] == "CVSS:4.0":
        version = "4.0"
    elif parts[0] == "CVSS:3.0":
        version = "3.0"
        
    if version in ("3.1", "3.0"):
        # Standard CVSS v3.1/3.0 Base Score calculation
        AV_map = {'N': 0.85, 'A': 0.62, 'L': 0.55, 'P': 0.2}
        AC_map = {'L': 0.77, 'H': 0.44}
        UI_map = {'N': 0.85, 'R': 0.62}
        PR_map = {
            'U': {'N': 0.85, 'L': 0.62, 'H': 0.27},
            'C': {'N': 0.85, 'L': 0.68, 'H': 0.5}
        }
        C_map = {'N': 0.0, 'L': 0.22, 'H': 0.56}
        I_map = {'N': 0.0, 'L': 0.22, 'H': 0.56}
        A_map = {'N': 0.0, 'L': 0.22, 'H': 0.56}
        
        AV = AV_map.get(metrics.get('AV'), 0.0)
        AC = AC_map.get(metrics.get('AC'), 0.0)
        UI = UI_map.get(metrics.get('UI'), 0.0)
        S = metrics.get('S', 'U')
        PR = PR_map.get(S, PR_map['U']).get(metrics.get('PR'), 0.0)
        C = C_map.get(metrics.get('C'), 0.0)
        I = I_map.get(metrics.get('I'), 0.0)
        A = A_map.get(metrics.get('A'), 0.0)
        
        exploitability = 8.22 * AV * AC * PR * UI
        iss = 1.0 - ((1.0 - C) * (1.0 - I) * (1.0 - A))
        if S == 'U':
            impact = 6.42 * iss
        else:
            impact = 7.52 * (iss - 0.029) - 3.25 * ((iss - 0.02) ** 15)
            
        if iss <= 0:
            return "0.0"
            
        int_val = int((exploitability + impact if S == 'U' else 1.08 * (exploitability + impact)) * 100000)
        score = (int_val / 100000.0) if int_val % 10000 == 0 else ((int(int_val / 10000) + 1) / 10.0)
        score = min(score, 10.0)
        score = max(score, 0.0)
        return f"{score:.1f}"
        
    elif version == "4.0":
        # CVSS v4.0 Scorer using Red Hat's Lookup Table
        def m(name):
            val = metrics.get(name, 'X')
            if name == 'E' and val == 'X': return 'A'
            if name in ('CR', 'IR', 'AR') and val == 'X': return 'H'
            return val
            
        # Determine eq1
        if m("AV") == "N" and m("PR") == "N" and m("UI") == "N":
            eq1 = "0"
        elif (
            (m("AV") == "N" or m("PR") == "N" or m("UI") == "N")
            and not (m("AV") == "N" and m("PR") == "N" and m("UI") == "N")
            and not m("AV") == "P"
        ):
            eq1 = "1"
        else:
            eq1 = "2"
            
        # Determine eq2
        if m("AC") == "L" and m("AT") == "N":
            eq2 = "0"
        else:
            eq2 = "1"
            
        # Determine eq3
        if m("VC") == "H" and m("VI") == "H":
            eq3 = "0"
        elif not (m("VC") == "H" and m("VI") == "H") and (m("VC") == "H" or m("VI") == "H" or m("VA") == "H"):
            eq3 = "1"
        else:
            eq3 = "2"
            
        # Determine eq4
        if m("MSI") == "S" or m("MSA") == "S":
            eq4 = "0"
        elif not (m("MSI") == "S" or m("MSA") == "S") and (m("SC") == "H" or m("SI") == "H" or m("SA") == "H"):
            eq4 = "1"
        else:
            eq4 = "2"
            
        # Determine eq5
        if m("E") == "A":
            eq5 = "0"
        elif m("E") == "P":
            eq5 = "1"
        else:
            eq5 = "2"
            
        # Determine eq6
        if (
            (m("CR") == "H" and m("VC") == "H")
            or (m("IR") == "H" and m("VI") == "H")
            or (m("AR") == "H" and m("VA") == "H")
        ):
            eq6 = "0"
        else:
            eq6 = "1"
            
        macro_vector = eq1 + eq2 + eq3 + eq4 + eq5 + eq6
        
        # Look up macro_vector in lookup table
        lookup_table = {
            "000000": 10.0, "000001": 9.9, "000010": 9.8, "000011": 9.5, "000020": 9.5, "000021": 9.2,
            "000100": 10.0, "000101": 9.6, "000110": 9.3, "000111": 8.7, "000120": 9.1, "000121": 8.1,
            "000200": 9.3, "000201": 9.0, "000210": 8.9, "000211": 8.0, "000220": 8.1, "000221": 6.8,
            "001000": 9.8, "001001": 9.5, "001010": 9.5, "001011": 9.2, "001020": 9.0, "001021": 8.4,
            "001100": 9.3, "001101": 9.2, "001110": 8.9, "001111": 8.1, "001120": 8.1, "001121": 6.5,
            "001200": 8.8, "001201": 8.0, "001210": 7.8, "001211": 7.0, "001220": 6.9, "001221": 4.8,
            "002001": 9.2, "002011": 8.2, "002021": 7.2, "002101": 7.9, "002111": 6.9, "002121": 5.0,
            "002201": 6.9, "002211": 5.5, "002221": 2.7,
            "010000": 9.9, "010001": 9.7, "010010": 9.5, "010011": 9.2, "010020": 9.2, "010021": 8.5,
            "010100": 9.5, "010101": 9.1, "010110": 9.0, "010111": 8.3, "010120": 8.4, "010121": 7.1,
            "010200": 9.2, "010201": 8.1, "010210": 8.2, "010211": 7.1, "010220": 7.2, "010221": 5.3,
            "011000": 9.5, "011001": 9.3, "011010": 9.2, "011011": 8.5, "011020": 8.5, "011021": 7.3,
            "011100": 9.2, "011101": 8.2, "011110": 8.0, "011111": 7.2, "011120": 7.0, "011121": 5.9,
            "011200": 8.4, "011201": 7.0, "011210": 7.1, "011211": 5.2, "011220": 5.0, "011221": 3.0,
            "012001": 8.6, "012011": 7.5, "012021": 5.2, "012101": 7.1, "012111": 5.2, "012121": 2.9,
            "012201": 6.3, "012211": 2.9, "012221": 1.7,
            "100000": 9.8, "100001": 9.5, "100010": 9.4, "100011": 8.7, "100020": 9.1, "100021": 8.1,
            "100100": 9.4, "100101": 8.9, "100110": 8.6, "100111": 7.4, "100120": 7.7, "100121": 6.4,
            "100200": 8.7, "100201": 7.5, "100210": 7.4, "100211": 6.3, "100220": 6.3, "100221": 4.9,
            "101000": 9.4, "101001": 8.9, "101010": 8.8, "101011": 7.7, "101020": 7.6, "101021": 6.7,
            "101100": 8.6, "101101": 7.6, "101110": 7.4, "101111": 5.8, "101120": 5.9, "101121": 5.0,
            "101200": 7.2, "101201": 5.7, "101210": 5.7, "101211": 5.2, "101220": 5.2, "101221": 2.5,
            "102001": 8.3, "102011": 7.0, "102021": 5.4, "102101": 6.5, "102111": 5.8, "102121": 2.6,
            "102201": 5.3, "102211": 2.1, "102221": 1.3,
            "110000": 9.5, "110001": 9.0, "110010": 8.8, "110011": 7.6, "110020": 7.6, "110021": 7.0,
            "110100": 9.0, "110101": 7.7, "110110": 7.5, "110111": 6.2, "110120": 6.1, "110121": 5.3,
            "110200": 7.7, "110201": 6.6, "110210": 6.8, "110211": 5.9, "110220": 5.2, "110221": 3.0,
            "111000": 8.9, "111001": 7.8, "111010": 7.6, "111011": 6.7, "111020": 6.2, "111021": 5.8,
            "111100": 7.4, "111101": 5.9, "111110": 5.7, "111111": 5.7, "111120": 4.7, "111121": 2.3,
            "111200": 6.1, "111201": 5.2, "111210": 5.7, "111211": 2.9, "111220": 2.4, "111221": 1.6,
            "112001": 7.1, "112011": 5.9, "112021": 3.0, "112101": 5.8, "112111": 2.6, "112121": 1.5,
            "112201": 2.3, "112211": 1.3, "112221": 0.6,
            "200000": 9.3, "200001": 8.7, "200010": 8.6, "200011": 7.2, "200020": 7.5, "200021": 5.8,
            "200100": 8.6, "200101": 7.4, "200110": 7.4, "200111": 6.1, "200120": 5.6, "200121": 3.4,
            "200200": 7.0, "200201": 5.4, "200210": 5.2, "200211": 4.0, "200220": 4.0, "200221": 2.2,
            "201000": 8.5, "201001": 7.5, "201010": 7.4, "201011": 5.5, "201020": 6.2, "201021": 5.1,
            "201100": 7.2, "201101": 5.7, "201110": 5.5, "201111": 4.1, "201120": 4.6, "201121": 1.9,
            "201200": 5.3, "201201": 3.6, "201210": 3.4, "201211": 1.9, "201220": 1.9, "201221": 0.8,
            "202001": 6.4, "202011": 5.1, "202021": 2.0, "202101": 4.7, "202111": 2.1, "202121": 1.1,
            "202201": 2.4, "202211": 0.9, "202221": 0.4,
            "210000": 8.8, "210001": 7.5, "210010": 7.3, "210011": 5.3, "210020": 6.0, "210021": 5.0,
            "210100": 7.3, "210101": 5.5, "210110": 5.9, "210111": 4.0, "210120": 4.1, "210121": 2.0,
            "210200": 5.4, "210201": 4.3, "210210": 4.5, "210211": 2.2, "210220": 2.0, "210221": 1.1,
            "211000": 7.5, "211001": 5.5, "211010": 5.8, "211011": 4.5, "211020": 4.0, "211021": 2.1,
            "211100": 6.1, "211101": 5.1, "211110": 4.8, "211111": 1.8, "211120": 2.0, "211121": 0.9,
            "211200": 4.6, "211201": 1.8, "211210": 1.7, "211211": 0.7, "211220": 0.8, "211221": 0.2,
            "212001": 5.3, "212011": 2.4, "212021": 1.4, "212101": 2.4, "212111": 1.2, "212121": 0.5,
            "212201": 1.0, "212211": 0.3, "212221": 0.1
        }
        score = lookup_table.get(macro_vector, 0.0)
        return f"{score:.1f}"
        
    return vector_str

def extract_fixed_version_from_advisory(item):
    """
    Statically extract the remediation (fixed) version for a vulnerability
    by inspecting its OSV affected ranges or NVD CPE match boundaries,
    ensuring the selected fix corresponds to the branch of the installed version.
    """
    import re
    from packaging import version as pkg_version
    from vuln_checker.feed_manager import safe_parse_version

    cve = item.get("cve", item)
    if not isinstance(cve, dict):
        return None
        
    # 1. Check if cve already has upgrade_to or fix_version
    # Note: If it matched an unrelated branch from a cached/global lookup, we re-evaluate below.
    # However, let's keep it as fallback if we don't have enough version info.
    target_version = item.get("component_version") or item.get("library_version")
    product_name = item.get("component_name") or item.get("library_name") or item.get("product")
    target_v = None
    if target_version and target_version != "N/A":
        try:
            target_v = safe_parse_version(target_version, product_name)
        except:
            pass

    # 2. Try OSV 'fixed' field if available
    affected_list = cve.get("affected", [])
    if isinstance(affected_list, list):
        osv_candidates = []
        for affected in affected_list:
            if not isinstance(affected, dict):
                continue
            for r in affected.get("ranges", []):
                if not isinstance(r, dict):
                    continue
                r_type = str(r.get("type", "")).upper()
                if r_type not in ("ECOSYSTEM", "SEMVER"):
                    continue # Skip GIT ranges as we cannot compare git commit hashes
                
                events = r.get("events", [])
                if not isinstance(events, list):
                    continue
                
                current_start = None
                intervals = []
                for ev in events:
                    if not isinstance(ev, dict):
                        continue
                    if "introduced" in ev:
                        current_start = ev["introduced"]
                    elif "fixed" in ev:
                        intervals.append({
                            "start": current_start or "0",
                            "end": ev["fixed"],
                            "inclusive": False
                        })
                        current_start = None
                    elif "last_affected" in ev:
                        intervals.append({
                            "start": current_start or "0",
                            "end": ev["last_affected"],
                            "inclusive": True
                        })
                        current_start = None
                
                # If target version is available, verify if it fits inside any interval
                matched_fix = None
                if target_v:
                    for interval in intervals:
                        try:
                            start_v = safe_parse_version(interval["start"], product_name) if interval["start"] != "0" else None
                            end_v = safe_parse_version(interval["end"], product_name) if interval["end"] else None
                            
                            lower_ok = True
                            upper_ok = True
                            
                            if start_v:
                                lower_ok = target_v >= start_v
                            if end_v:
                                if interval["inclusive"]:
                                    upper_ok = target_v <= end_v
                                else:
                                    upper_ok = target_v < end_v
                                    
                            if lower_ok and upper_ok:
                                matched_fix = interval["end"] if not interval["inclusive"] else f"> {interval['end']}"
                                break
                        except:
                            pass
                else:
                    # Take the last interval's fixed version as fallback
                    if intervals:
                        last_int = intervals[-1]
                        matched_fix = last_int["end"] if not last_int["inclusive"] else f"> {last_int['end']}"
                
                if matched_fix:
                    priority = 3 if "ECOSYSTEM" in r_type else 2
                    osv_candidates.append((priority, matched_fix))
                    
        if osv_candidates:
            osv_candidates.sort(key=lambda x: x[0], reverse=True)
            return osv_candidates[0][1]

    # 3. Try NVD CPE match boundaries (versionEndExcluding / versionEndIncluding)
    configurations = cve.get("configurations", [])
    if isinstance(configurations, list):
        for config in configurations:
            if not isinstance(config, dict):
                continue
            for node in config.get("nodes", []):
                if not isinstance(node, dict):
                    continue
                for cpe_match in node.get("cpeMatch", []):
                    if not isinstance(cpe_match, dict):
                        continue
                    if cpe_match.get("vulnerable"):
                        criteria = cpe_match.get("criteria", "")
                        cpe_used = item.get("cpe_used")
                        if criteria and cpe_used:
                            from vuln_checker.utils import get_vendor_aliases
                            crit_parts = criteria.split(":")
                            used_parts = cpe_used.split(":")
                            if len(crit_parts) >= 5 and len(used_parts) >= 5:
                                crit_vendor = crit_parts[3].lower()
                                crit_product = crit_parts[4].lower()
                                used_vendor = used_parts[3].lower()
                                used_product = used_parts[4].lower()
                                
                                # 1. Check vendor mismatch (allowing aliases)
                                if crit_vendor != used_vendor:
                                    aliases_config = get_vendor_aliases()
                                    aliases = aliases_config.get("aliases", {})
                                    is_alias = False
                                    for v, v_aliases in aliases.items():
                                        v_lower = v.lower()
                                        v_aliases_lower = [a.lower() for a in v_aliases]
                                        if (crit_vendor == v_lower or crit_vendor in v_aliases_lower) and \
                                           (used_vendor == v_lower or used_vendor in v_aliases_lower):
                                            is_alias = True
                                            break
                                    if not is_alias:
                                        continue
                                        
                                # 2. Check product mismatch
                                if crit_product != used_product:
                                    if not (used_product.startswith(crit_product) or crit_product.startswith(used_product)):
                                        continue

                        v_start = cpe_match.get("versionStartIncluding")
                        v_start_ex = cpe_match.get("versionStartExcluding")
                        v_end = cpe_match.get("versionEndIncluding")
                        v_end_ex = cpe_match.get("versionEndExcluding")
                        
                        # Verify if target version fits in this specific CPE match boundary
                        if target_v:
                            try:
                                lower_match = True
                                upper_match = True
                                if v_start:
                                    lower_match = target_v >= safe_parse_version(v_start, product_name)
                                elif v_start_ex:
                                    lower_match = target_v > safe_parse_version(v_start_ex, product_name)
                                if v_end:
                                    upper_match = target_v <= safe_parse_version(v_end, product_name)
                                elif v_end_ex:
                                    upper_match = target_v < safe_parse_version(v_end_ex, product_name)
                                    
                                if not (lower_match and upper_match):
                                    continue
                            except:
                                pass
                        
                        if v_end_ex:
                            return v_end_ex
                        if v_end:
                            return f"> {v_end}"

    # Fallback to the first available non-empty upgrade_to/fix_version
    for key in ["upgrade_to", "fix_version"]:
        if item.get(key) and item.get(key) != "N/A":
            return item.get(key)

    return None

def _extract_cve_data(item):
    """Unified field extractor for any CVE finding item with robust metric parsing."""
    # Check if item is already extracted (has our custom keys)
    if "cvssScore" in item and "url" in item:
        return item
        
    cve = item.get("cve", item)
    
    # Severity & Score extraction using prioritized helper
    from vuln_checker.utils import extract_cve_severity_and_score
    severity, base_score = extract_cve_severity_and_score(cve)
    score = calculate_cvss_score(base_score)


    # Component Name & Version logic (with fallback)
    comp_name = item.get("component_name")
    lib_ver = item.get("component_version")
    if not comp_name or not lib_ver or lib_ver == "N/A":
        product_full = item.get("product", "Unknown")
        # Remove parent jar info if present: "group:name:version (parent.jar)"
        temp_name = product_full.split(" (")[0]
        parts = temp_name.split(":")
        if len(parts) >= 2:
            if not lib_ver or lib_ver == "N/A":
                lib_ver = parts[-1]
            if not comp_name or comp_name == "N/A":
                comp_name = ":".join(parts[:-1])
        else:
            comp_name = comp_name or temp_name
            lib_ver = lib_ver or "N/A"

    orig_id = cve.get("id", "N/A")
    cve_id = orig_id
    url = ""
    
    # Extract aliases
    aliases = item.get("aliases", []) or cve.get("aliases", [])
    cve_alias = next((a for a in aliases if str(a).startswith("CVE-")), None)
    ghsa_alias = next((a for a in aliases if str(a).startswith("GHSA-")), None)

    # Logic for ID and URL
    if str(orig_id).startswith("RETIRE-"):
        cve_id = ""
        url = ""
    elif item.get("url"):
        url = item.get("url")
    elif str(orig_id).startswith("GHSA-"):
        url = f"https://github.com/advisories/{orig_id}"
    elif str(orig_id).startswith("CVE-"):
        url = f"https://nvd.nist.gov/vuln/detail/{orig_id}"
    elif ghsa_alias:
        url = f"https://github.com/advisories/{ghsa_alias}"
    elif cve_alias:
        url = f"https://nvd.nist.gov/vuln/detail/{cve_alias}"
    elif orig_id != "N/A":
        url = f"https://nvd.nist.gov/vuln/detail/{orig_id}"
    
    return {
        "id": cve_id,
        "severity": severity.upper() if severity else "N/A",
        "cvssScore": score,
        "component_name": comp_name,
        "component_version": lib_ver,
        "path": item.get("path", "N/A"),
        "purl": item.get("purl", "N/A"),
        "cpe_used": item.get("cpe_used", "N/A"),
        "cpe_source": item.get("cpe_source", "N/A"),
        "feed_file": item.get("feed_file") or item.get("nvd_feed_file") or cve.get("feed_file") or cve.get("nvd_feed_file", "unknown"),
        "published": cve.get("published", "N/A"),
        "lastModified": cve.get("lastModified", "N/A"),
        "description": cve.get("descriptions", [{}])[0].get("value", "N/A") if isinstance(cve.get("descriptions"), list) else cve.get("details", "N/A"),
        "dev_comments": "",
        "url": url,
        "upgrade_to": extract_fixed_version_from_advisory(item) or item.get("upgrade_to") or item.get("fix_version") or "N/A",
        "delta_status": item.get("delta_status", ""),
        "aliases": aliases,
        "pkg_type": item.get("pkg_type") or cve.get("pkg_type") or "",
        "tool": item.get("tool") or "unknown"
    }

def output_results(cves, output_format="json", output_file=None, has_baseline=False):
    if not cves:
        logger.warning("⚠️ No CVEs found.")
        return

    processed_data = [_extract_cve_data(item) for item in cves]
    
    if output_format == "json":
        target = output_file or "cve_report.json"
        if not has_baseline:
            processed_data = [{k: v for k, v in row.items() if k != "delta_status"} for row in processed_data]
        with open(target, "w", encoding="utf-8") as f:
            json.dump(processed_data, f, indent=2)
        logger.info(f"✅ JSON report written to {target}")
        
    elif output_format == "csv":
        target = output_file or "cve_report.csv"
        active_columns = get_report_columns(has_baseline)
        headers = [col[1] for col in active_columns]
        keys = [col[0] for col in active_columns]
        
        with open(target, "w", newline='', encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=keys, extrasaction='ignore')
            writer.writerow(dict(zip(keys, headers)))
            for row in processed_data:
                # Truncate description for CSV to avoid massive cells
                desc = row.get("description", "")
                if len(desc) > 500:
                    desc = desc[:500] + "..."
                
                # Excel/CSV hyperlink for IDs
                if row.get("id") and row.get("url"):
                    writer.writerow({**row, "description": desc, "id": f'=HYPERLINK("{row["url"]}", "{row["id"]}")'})
                else:
                    writer.writerow({**row, "description": desc})
        logger.info(f"✅ CSV report written to {target}")

def output_results_sbom_enhanced(cves, output_format="json", output_file=None, has_baseline=False):
    """Fallback for direct call from main.py if needed, using unified logic."""
    output_results(cves, output_format, output_file, has_baseline)

def generate_html_report(cves, output_file="cve_report.html", has_baseline=False):
    script_dir = Path(__file__).parent
    template_dir = script_dir / "templates"
    env = Environment(
        loader=FileSystemLoader(str(template_dir)),
        autoescape=select_autoescape(['html', 'xml'])
    )
    try:
        template = env.get_template("template.html")
    except Exception as e:
        logger.error(f"❌ Error loading template.html: {e}")
        return

    processed_data = [_extract_cve_data(item) for item in cves]
    categories = categorize_findings(processed_data)
    
    severity_counter = Counter([row["severity"] for row in processed_data])
    
    category_order = ["Jars / War / Zip", "Npm audit", "RetireJS", "Pip-audit", "OS & System Packages"]
    present_categories = []
    
    severity_order = {
        "CRITICAL": 4,
        "HIGH": 3,
        "MEDIUM": 2,
        "LOW": 1,
        "N/A": 0,
        "UNKNOWN": 0
    }
    
    for cat_name in category_order:
        if cat_name in categories:
            rows = categories[cat_name]
            rows.sort(key=lambda r: severity_order.get(r.get("severity", "UNKNOWN").upper(), 0), reverse=True)
            present_categories.append({
                "name": cat_name,
                "safe_id": cat_name.lower().replace(" / ", "_").replace(" ", "_").replace("&", "and").replace("-", "_"),
                "rows": rows
            })
            
    from datetime import datetime
    html = template.render(
        categories=present_categories,
        severity_counts=severity_counter,
        now=datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        columns=get_report_columns(has_baseline)
    )
    with open(output_file, "w", encoding="utf-8") as f:
        f.write(html)
    logger.info(f"📄 HTML report written to {output_file}")

def output_results_excel(cves, output_file="cve_report.xlsx", has_baseline=False, packages_info=None, config=None):
    wb = openpyxl.Workbook()
    
    processed_data = [_extract_cve_data(item) for item in cves]
    categories = categorize_findings(processed_data)
    
    if not categories:
        # If absolutely no findings, create a single empty Libraries sheet
        ws = wb.active
        ws.title = "Libraries"
        active_columns = get_report_columns(has_baseline)
        headers = [col[1] for col in active_columns]
        ws.append(headers)
        wb.save(output_file)
        logger.info(f"✅ Empty Excel report written to {output_file}")
        return
        
    severity_order = {
        "CRITICAL": 4,
        "HIGH": 3,
        "MEDIUM": 2,
        "LOW": 1,
        "N/A": 0,
        "UNKNOWN": 0
    }
    
    active_columns = get_report_columns(has_baseline)
    headers = [col[1] for col in active_columns]
    
    first_sheet = True
    category_order = ["Jars / War / Zip", "Npm audit", "RetireJS", "Pip-audit", "OS & System Packages"]
    present_categories = [cat for cat in category_order if cat in categories]
    
    for cat_name in present_categories:
        rows = categories[cat_name]
        rows.sort(key=lambda r: severity_order.get(r.get("severity", "UNKNOWN").upper(), 0), reverse=True)
        
        tab_title = cat_name.replace("/", "-").replace("\\", "-").replace("?", "").replace("*", "").replace(":", "-").replace("[", "").replace("]", "")
        if len(tab_title) > 31:
            tab_title = tab_title[:31]
            
        if first_sheet:
            ws = wb.active
            ws.title = tab_title
            first_sheet = False
        else:
            ws = wb.create_sheet(title=tab_title)
            
        ws.append(headers)
        for col_num in range(1, len(headers) + 1):
            ws.cell(row=1, column=col_num).font = Font(bold=True)
            
        def clean_val(v):
            import re
            if v is None:
                return ""
            if not isinstance(v, str):
                v = str(v)
            return re.sub(r'[\000-\010\013\014\016-\037]', '', v)

        for idx, row in enumerate(rows, start=2):
            for col_idx, (key, _) in enumerate(active_columns, start=1):
                val = clean_val(row.get(key, ""))
                cell = ws.cell(row=idx, column=col_idx, value=val)
                if key == "id" and val and val != "N/A" and row.get("url"):
                    cell.hyperlink = row.get("url")
                    cell.font = Font(color="0000FF", underline="single")
                elif key == "path":
                    cell.alignment = Alignment(wrap_text=True)
                    
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"
        
    # Add License Summary Sheet if packages_info is available
    if packages_info:
        from vuln_checker.utils import check_license_compliance
        ws_lic = wb.create_sheet(title="License Summary")
        lic_headers = ["Vendor", "Product", "Version", "License", "Type", "Compliance Status", "Path"]
        ws_lic.append(lic_headers)
        for col_num in range(1, len(lic_headers) + 1):
            ws_lic.cell(row=1, column=col_num).font = Font(bold=True)
            
        row_idx = 2
        for key, entries in packages_info.items():
            best_entry = None
            for entry in entries:
                if not best_entry:
                    best_entry = entry
                    continue
                current_best_lic = best_entry.get("license", "unknown")
                new_lic = entry.get("license", "unknown")
                if current_best_lic == "unknown" and new_lic != "unknown":
                    best_entry = entry
                elif new_lic != "unknown" and len(new_lic) > len(current_best_lic):
                    best_entry = entry
            
            if best_entry:
                license_str = best_entry.get("license", "unknown")
                status = best_entry.get("license_status")
                if not status:
                    status = check_license_compliance(license_str, config or {})
                
                ws_lic.append([
                    clean_val(best_entry.get("vendor", "")),
                    clean_val(best_entry.get("name", "")),
                    clean_val(best_entry.get("version", "")),
                    clean_val(license_str),
                    clean_val(best_entry.get("type", "")),
                    clean_val(status),
                    clean_val(best_entry.get("path", ""))
                ])
                status_cell = ws_lic.cell(row=row_idx, column=6)
                if status == "VIOLATED":
                    status_cell.font = Font(color="FF0000", bold=True)
                elif status == "REVIEW":
                    status_cell.font = Font(color="FFA500", bold=True)
                elif status == "PASS":
                    status_cell.font = Font(color="008000", bold=True)
                row_idx += 1
                
        if row_idx > 2:
            ws_lic.auto_filter.ref = f"A1:G{row_idx - 1}"
        
    wb.save(output_file)
    logger.info(f"✅ Excel report written to {output_file}")
